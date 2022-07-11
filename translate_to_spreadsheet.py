#This will NOT work on teton at the moment because it does not have openpyxl
import re
import os

#this got auto added? idk what it is

#openpyxl should let us edit and access excel files
from openpyxl import Workbook
from openpyxl import load_workbook



#this way so we can iterate over fields
fields={"Location: ":"F", "Material Dated: ":"C",
"Lab Name: ":"AF","Lab Number: ":"A","Age: ":"Q","Age Sigma: ":"R","Latitude: ":"G","Longitude: ":"H","Type Of Date: ":"E"}

#dicts to keep track of our materials
materials={}


#These are all of the
#  possible fields
location = ""
materialDated = ""
labName = ""
labNumber = ""
age = ""
ageSigma = ""
latitude = ""
longitude = ""
typeOfDate = ""
siteIdentifier = ""

ageAssigned = 0


#Where our data will come from and go to
input = 'D:\\ARCC\\radiocarbon\\radiocarbonOutputToSpreadsheet\\newfullinput'
output = 'D:\\ARCC\\radiocarbon\\radiocarbonOutputToSpreadsheet\\output'

def translate_to_spreadsheet(dirinput):

    wb = load_workbook('CARD Upload Template.xlsx')
    ws =wb['Data Fields']
    #counters to know what row and column we are at
    #col='A'
    spreadsheet_final_row=5
    errorcount=0
    folders=os.listdir(dirinput)
    for folder in folders:
        print(folder)
        input = os.path.join(dirinput,folder)
        for file in os.listdir(input):
                        
            #obtain the file number and store it in filenum, then put that in our spreadsheet
            filenum =file.split("_")[0]
            ws["AC"+str(spreadsheet_final_row)]=str(filenum)

            #open the file and iterate over its lines
            with open(os.path.join(input,file), "r",encoding="utf8") as text_file:
                for line in text_file:
                    #use this for looping. doing it this way so i can easily access index of strings in fields
                    
                    for key in fields:
                        #check if each field is found on this line
                        field_found=re.search(key, line)
                        if field_found:
                            #make sure the key has an entry
                            if fields[key]:
                                #remove the title and then the first space from the line and also remove any newline characters
                                content=line.split(":")[1]
                                content=content[1:]
                                content=content.strip('\r\n')
                                content=content.strip('\n')
                                #65=A
                                col=fields[key]
                                try:
                                    ws[col+str(spreadsheet_final_row)]=str(content)
                                except:
                                    errorcount=errorcount+1


                            #we can also break once the field is found since there is only one per line
                            break
                        
            #using ascii we can increment columns
            #col=chr(ord(col) + 3)

            #increment to the next entry for the next card
            spreadsheet_final_row=spreadsheet_final_row+1


    #finally we save the spreadsheet
    wb.save('Card_Upload_Template_Output.xlsx')
    print(errorcount," errors")



translate_to_spreadsheet(input)

