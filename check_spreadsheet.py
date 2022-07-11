#check_spreadsheet.py contains functions used to verify and in some cases correct data entries in the excel file

#This will NOT work on teton at the moment because it does not have openpyxl
from asyncio.windows_events import NULL
import re
import os

#openpyxl should let us edit and access excel files
from openpyxl import Workbook
from openpyxl import load_workbook
#so we can mark cells
from openpyxl.styles import Font
red = Font(color="FF0000")

#this way we can iterate over fields
fields={"Location: ":"F", "Material Dated: ":"C",
"Lab Name: ":"AF","Lab Number: ":"A","Age: ":"Q","Age Sigma: ":"R","Latitude: ":"G","Longitude: ":"H","Type Of Date: ":"E"}

#here is a place to store the materials when we use catalogMAterialDated
materials={}

#this function will check an entry for errors, checking fields from left to right. If an error is found it does not check the other fields
def checkEntry(row,worksheetIn):
    print(row)    
    #assume the entry is good until an error is found
    good_entry=True
    column=NULL
    #if statements for each implemented field check
    if not checkLab_Number(row,worksheetIn):
        good_entry=False
        column='A'
        return [good_entry,column]
    
    if not checkMaterial(row,worksheetIn): 
        good_entry=False
        column='C'
        return [good_entry,column]
    
    return [good_entry,column]

#this function takes a row, finds its material and catlogs in in a text document
#this way we can see all of the materials that are being read in
def catalogMaterial_Dated(row,worksheetIn):
    
    text=worksheetIn['C'+str(row)].value

    #in order for case insensitive comparison, we also need to make sure that it actually has a material field and if not, record that
    try:
        lowerStr=text.lower()
        lowerStr=lowerStr.replace(" ","")
        lowerStr=lowerStr.replace("-","")
        lowerStr=lowerStr.replace("_","")
        lowerStr=lowerStr.replace("+","")
        lowerStr=lowerStr.replace("^","")
        lowerStr=lowerStr.replace("•","")
        lowerStr=lowerStr.replace(",","")
        lowerStr=lowerStr.replace("[","")
        lowerStr=lowerStr.replace("]","")
        lowerStr=lowerStr.replace("(","")
        lowerStr=lowerStr.replace(")","")
        lowerStr=lowerStr.replace(".","")
        
    #some entries will have strange characters so we catch them here
    except:
        lowerStr='materialfielderror'
        

    #these are the materials we have found so far
    mats=materials.keys()
    #iterate through keys and compare without regard for case, keep a tally of each material
    found=False
    for key in mats:
        if key==lowerStr:
            materials[key]=materials[key]+1
            found=True
            break
    if not found:
        materials[lowerStr]=1
    return

#this basically just calls checkentries interatively to check the whole document
def checkAllEntries(worksheet,start,end):
    
    #load new worksheet to store bad and good entries
    bad_wb=load_workbook('CARD Upload Template.xlsx')
    bad_ws =bad_wb['Data Fields']
    good_wb=load_workbook('CARD Upload Template.xlsx')
    good_ws =good_wb['Data Fields']
    
    #iterate over the rows in our range and check them, if they are bad, add them to a the spreadsheet
    #row counter to pass to checkentry
    rowcounter=start
    #finding the max_row of the worksheet is very slow so we do this instead
    bad_output_current_row=5
    good_output_current_row=5
    
    for row in worksheet.iter_rows(min_row=start, max_col=33, max_row=end):
        catalogMaterial_Dated(rowcounter,worksheet)
        output=checkEntry(rowcounter,worksheet)
        if not output[0]:     
            #if there is a bad row we copy all of its cells into our new spreadsheet       
            for cell in row:
                #for some reason it only works if we assign the address seperately like this
                addy=str(cell.column_letter)+str(bad_output_current_row)
                v=cell.value
                #assign the value to the corresponding cell
                bad_ws[addy]= v

            #change text to red so we can see what error was found
            bad_ws[str(output[1])+str(bad_output_current_row)].font=red
            bad_output_current_row=bad_output_current_row+1
        else:
            
            for cell in row:
                #for some reason it only works if we assign the address seperately liek thi
                addy=str(cell.column_letter)+str(good_output_current_row)
                v=cell.value
                #assign the value to the corresponding cell
                good_ws[addy]= v
            good_output_current_row=good_output_current_row+1
        rowcounter=rowcounter+1
            
    
    bad_wb.save("Bad_Cards.xlsx")
    good_wb.save("Good_Cards.xlsx")
    return

#will check the lab number of an entry
#this function may also attempt to correct the entry and change it
def checkLab_Number(row,worksheet):
    
    #lab numbers must be in AAA-nnn with varying numbers of characters according to the instructions of the Canadian radiocarbon database
    accepted_forms=["[A-Za-z]-[0-9][0-9][0-9][0-9]" #A-nnnn
                    ,"[A-Za-z]-[0-9][0-9][0-9]"#A-nnn
                    ,"[A-Za-z]-[0-9][0-9]"#A-nn
                    ,"[A-Za-z]-[0-9]"#A-n
                    ,"[A-Za-z][A-Za-z]-[0-9][0-9][0-9][0-9]"#AA-nnnn
                    ,"[A-Za-z][A-Za-z]-[0-9][0-9][0-9]"#AA-nnn
                    ,"[A-Za-z][A-Za-z]-[0-9][0-9]"#AA-nn
                    ,"[A-Za-z][A-Za-z]-[0-9]"#AA-n
                    ,"[A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9][0-9][0-9]"#AAA-nnnn
                    ,"[A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9][0-9]"#AAA-nnn
                    ,"[A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9]"#AAA-nn
                    ,"[A-Za-z][A-Za-z][A-Za-z]-[0-9]"#AAA-n
                    ,"[A-Za-z][A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9][0-9][0-9]"#AAAA-nnnn
                    ,"[A-Za-z][A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9][0-9]"#AAAA-nnn
                    ,"[A-Za-z][A-Za-z][A-Za-z][A-Za-z]-[0-9][0-9]"#AAAA-nn
                    ,"[A-Za-z][A-Za-z][A-Za-z][A-Za-z]-[0-9]"#AAAA-n
                    ]
    text=worksheet['A'+str(row)].value
    
    #use found to track if text is in accepted form
    
    found=False
    #cycle through accepted forms, if it is in an accepted form then we are done. if not, then found=False and we move on to check for any matches in the text
    # , not just exact matches
    for regex in accepted_forms:
        try:
            match=re.fullmatch(regex,text)
        except:
            break
        if match !=None:
            found =re.fullmatch(regex,text).group(0)
            break

    #if we didnt find an exact match, we start looking for parts of the sring that match, and take the longest of these as the match
    
    if not found:
        potential_match=""
        for regex in accepted_forms:        
            try:
                search=re.search(regex,text)

                
                if len(search.group(0))>len(potential_match):
                    potential_match=search.group(0)
            except:
                pass

        if len(potential_match)>0:
            worksheet['A'+str(row)].value=potential_match
            found=potential_match
            #keep track of what we extracted from its larger string
            extractedLabNumbers.write("Card "+worksheet['AC'+str(row)].value+" : "+found+ " out of "+text+os.linesep)  
    
    #then if that didnt work, remove spaces and look for matches
    if not found:
        try:
            newtext=text.replace(" ","")
            potential_match=""
            for regex in accepted_forms:        
                try:
                    search=re.search(regex,newtext)

                    
                    if len(search.group(0))>len(potential_match):
                        potential_match=search.group(0)
                except:
                    pass

            if len(potential_match)>0:
                worksheet['A'+str(row)].value=potential_match
                found=potential_match
                #keep track of what we extracted from its larger string
                extractedLabNumbers.write("Card "+worksheet['AC'+str(row)].value+" : "+found+ " out of "+text+os.linesep) 
        except:
            pass 
    
    return found
#sorts and lists the materials in the materials dict to a text file
def listMaterials():

    file = open("DictFile.txt","w",encoding="utf-8")
    
    for w in sorted(materials, key=materials.get, reverse=True):
        file.write('%s:%s\n' % (w, materials[w]))
    
    file.close()
    return

#reads in a dict of valid materials from a text file
def getValidMaterials(filename):
    materialsFile=open(filename,encoding="utf-8")
    validMaterials={}
    
    #we get rid of the tally
    for line in materialsFile.readlines():
        text=line.split(':')[0]
        validMaterials[text]=True
    return validMaterials

#references a material against the valid materials to check its validity
def checkMaterial(row, worksheet):
    text=worksheet['C'+str(row)].value
    
    try:
        text=text.lower()
        text=text.replace(" ","")
        text=text.replace("-","")
        text=text.replace("_","")
        text=text.replace("+","")
        text=text.replace("^","")
        text=text.replace("•","")
        text=text.replace(",","")
        text=text.replace("[","")
        text=text.replace("]","")
        text=text.replace("(","")
        text=text.replace(")","")
        text=text.replace(".","")
    except:
        return False
    found=False

    try:
        if validMaterials[text]:
            found=text
    except:
        found=False

    return found

#this will use the materials and crossrefererences with the valid materials to find what I manually removed from validmaterials
def generateRemoved():
    removedMats = open("Removed_materials.txt","w",encoding="utf-8")
    for key in materials.keys():
        try:
            x=validMaterials[key]
        except:
            removedMats.write(key+os.linesep)
    removedMats.close()



wb=load_workbook('Card_Upload_Template_Output.xlsx')
ws =wb['Data Fields']
extractedLabNumbers = open("Extracted_Lab_Numbers.txt","w",encoding="utf-8")

validMaterials=getValidMaterials("Valid Materials.txt")




checkAllEntries(ws,5,ws.max_row)


extractedLabNumbers.close()
listMaterials()


print('done')

